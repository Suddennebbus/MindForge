from pathlib import Path
from typing import Optional

# PDF metadata fields worth surfacing to the LLM during ingest. The arXivID /
# DOI / Author live in file properties, not the extracted text (page footers
# get garbled), which otherwise leads the model to guess identifiers.
_METADATA_KEYS = ("Title", "Author", "Subject", "Keywords", "DOI", "arXivID")


def parse_document_metadata(file_path: str) -> dict:
    """Extract trustworthy document properties (currently PDF only)."""
    path = Path(file_path)
    if path.suffix.lower() != ".pdf":
        return {}
    import pdfplumber
    try:
        with pdfplumber.open(path) as pdf:
            raw = pdf.metadata or {}
    except Exception:
        return {}
    return {k: str(raw[k]) for k in _METADATA_KEYS if raw.get(k)}


def parse_document(file_path: str) -> str:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        return _parse_pdf(path)
    elif suffix == ".docx":
        return _parse_docx(path)
    elif suffix == ".xlsx":
        return _parse_xlsx(path)
    elif suffix in (".html", ".htm"):
        return _parse_html(path)
    elif suffix in (".txt", ".md"):
        return path.read_text(encoding="utf-8")
    else:
        return path.read_text(encoding="utf-8", errors="ignore")


def _parse_pdf(path: Path) -> str:
    import pdfplumber
    text_parts = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
    return "\n\n".join(text_parts)


def _parse_docx(path: Path) -> str:
    from docx import Document
    doc = Document(path)
    return "\n\n".join([p.text for p in doc.paragraphs if p.text.strip()])


def _parse_xlsx(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                parts.append(row_text)
    return "\n".join(parts)


def _parse_html(path: Path) -> str:
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    for script in soup(["script", "style"]):
        script.decompose()
    return soup.get_text(separator="\n", strip=True)
